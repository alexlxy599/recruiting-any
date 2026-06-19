"""ICML 总表 Medium 置信度候选人的快速信号复核（第一段，不用 LLM）。

对每个 GH置信度=medium 的人跑四个零推理信号:
  1. GitHub profile 公开邮箱 / commit 邮箱 == 表内邮箱      → 升高置信
  2. GitHub blog 字段 == 表内个人主页（同一站点）           → 升高置信
  3. 表内 LinkedIn slug 出现在 bio/blog/README/主页 HTML    → 升高置信
  4. 姓名拼音 ↔ login 吻合 且 机构关键词出现在 GitHub 证据  → 升高置信
  profile 404                                               → 排除(死链)
  其余维持 medium，证据存档供第二段 LLM 仲裁。

结果写 data/medium_review.jsonl（断点续传），跑完后用 --writeback 回写总表新列"GH复核"。
"""

import argparse
import json
import os
import re
import sys
import time
import urllib.parse

import requests
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from verify_github import (gh_headers, fetch_profile, fetch_profile_readme,
                           fetch_homepage, fetch_commit_emails, load_env)
from ingest.import_icml import login_matches_name

load_env()

XLSX = "/Users/alex/Desktop/ICML 2026 总表0611.xlsx"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "medium_review.jsonl")

STOP_WORDS = {"university", "institute", "of", "the", "and", "for", "college",
              "school", "department", "lab", "laboratory", "research", "center",
              "centre", "science", "sciences", "technology", "chinese", "national"}


def norm_site(u: str) -> str:
    u = (u or "").strip().lower()
    u = re.sub(r"^https?://", "", u).rstrip("/")
    return re.sub(r"^www\.", "", u)


def inst_keywords(inst: str) -> list[str]:
    words = re.findall(r"[a-z]{4,}", (inst or "").lower())
    return [w for w in words if w not in STOP_WORDS][:3]


def load_medium() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb["全部华人作者"]
    rows = ws.iter_rows(values_only=True)
    idx = {h: i for i, h in enumerate(next(rows))}
    people = {}
    for r in rows:
        def g(col):
            v = r[idx[col]]
            return str(v).strip() if v is not None else ""
        if g("GH置信度").lower() != "medium":
            continue
        gh = g("GitHub").lower().rstrip("/")
        m = re.search(r"github\.com/([^/?#]+)", gh)
        if not m or gh in people:
            continue
        people[gh] = {
            "login": m.group(1),
            "github": gh,
            "author": g("Author"),
            "email": (g("Email") or g("GH Email")).lower(),
            "homepage": g("个人主页"),
            "linkedin": g("LinkedIn"),
            "institution": g("Institution") or g("Org Label"),
        }
    wb.close()
    return list(people.values())


def check_person(p: dict) -> dict:
    login = p["login"]
    profile, remaining = fetch_profile(login)
    if profile is None:
        return {"verdict": "SKIPPED", "reason": "api_error", "remaining": remaining}
    if profile == {}:
        return {"verdict": "排除", "reason": "GitHub 账号不存在", "remaining": remaining}

    bio = profile.get("bio") or ""
    blog = profile.get("blog") or ""
    gh_name = (profile.get("name") or "").strip()
    company = profile.get("company") or ""

    # 信号 1a: profile 公开邮箱
    if p["email"] and (profile.get("email") or "").lower() == p["email"]:
        return {"verdict": "升高置信", "reason": f"profile邮箱={p['email']}", "remaining": remaining}

    # 信号 2: blog == 表内个人主页
    if p["homepage"] and norm_site(blog) and norm_site(blog) == norm_site(p["homepage"]):
        return {"verdict": "升高置信", "reason": f"GitHub挂的主页与表内一致 {norm_site(blog)}",
                "remaining": remaining}

    readme = fetch_profile_readme(login)
    homepage_html, homepage_text = "", ""
    hp_url = p["homepage"] or (blog if blog and "github.com" not in blog else "")
    if hp_url:
        if not hp_url.startswith("http"):
            hp_url = "https://" + hp_url
        homepage_html, homepage_text = fetch_homepage(hp_url)

    evidence = " ".join([bio, blog, company, gh_name, readme, homepage_html]).lower()

    # 信号 3: LinkedIn slug 反链
    if p["linkedin"]:
        m = re.search(r"linkedin\.com/in/([^/?#]+)", p["linkedin"], re.I)
        slug = urllib.parse.unquote(m.group(1)).lower().rstrip("/") if m else ""
        if slug and slug in evidence:
            return {"verdict": "升高置信", "reason": f"LinkedIn slug '{slug}' 反链", "remaining": remaining}

    # 信号 1b: commit 邮箱（多 1 次 API 调用，只在有邮箱时做）
    if p["email"]:
        if p["email"] in fetch_commit_emails(login):
            return {"verdict": "升高置信", "reason": f"commit邮箱={p['email']}", "remaining": remaining}

    # 信号 4: 姓名 ↔ login/profile名 吻合 且 机构词出现在证据里
    name_ok = login_matches_name(login, p["author"]) or \
        (gh_name and gh_name.lower() == p["author"].lower())
    kws = inst_keywords(p["institution"])
    inst_ok = any(k in evidence for k in kws)
    if name_ok and inst_ok:
        return {"verdict": "升高置信", "reason": f"姓名吻合+机构词命中({[k for k in kws if k in evidence][0]})",
                "remaining": remaining}

    return {"verdict": "维持", "reason": "无决定性信号",
            "evidence": {"bio": bio[:300], "name": gh_name, "company": company,
                         "blog": blog, "readme": readme[:1500], "homepage": homepage_text[:2000]},
            "remaining": remaining}


def run(workers: int = 10):
    import threading
    from concurrent.futures import ThreadPoolExecutor, as_completed

    people = load_medium()
    done = set()
    if os.path.exists(OUT):
        for line in open(OUT, encoding="utf-8"):
            try:
                done.add(json.loads(line)["github"])
            except Exception:
                pass
    todo = [p for p in people if p["github"] not in done]
    print(f"medium 共 {len(people)} 人, 已处理 {len(done)}, 本次 {len(todo)}, 并发 {workers}", flush=True)

    f = open(OUT, "a", encoding="utf-8")
    lock = threading.Lock()
    stats = {}
    n_done = 0

    def work(p):
        try:
            return p, check_person(p)
        except Exception as e:
            return p, {"verdict": "SKIPPED", "reason": f"error: {e}", "remaining": -1}

    with ThreadPoolExecutor(max_workers=workers) as ex:
        futures = [ex.submit(work, p) for p in todo]
        for fut in as_completed(futures):
            p, r = fut.result()
            with lock:
                if r["verdict"] != "SKIPPED" or r.get("remaining") != 0:
                    rec = {**{k: p[k] for k in ("github", "login", "author", "institution")}, **r}
                    f.write(json.dumps(rec, ensure_ascii=False) + "\n")
                    f.flush()
                stats[r["verdict"]] = stats.get(r["verdict"], 0) + 1
                n_done += 1
                if n_done % 100 == 0:
                    print(f"[{n_done}/{len(todo)}] {stats}", flush=True)
                rem = r.get("remaining", 5000)
            if 0 < rem < 300:
                time.sleep(20)  # 接近限流时全体放缓
    f.close()
    print("完成:", stats, flush=True)


def writeback():
    results = {}
    for line in open(OUT, encoding="utf-8"):
        r = json.loads(line)
        if r["verdict"] in ("升高置信", "排除"):
            results[r["github"]] = r["verdict"]
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["全部华人作者"]
    header = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(header)}
    col = idx.get("GH复核")
    if not col:
        col = len(header) + 1
        ws.cell(row=1, column=col, value="GH复核")
    n = 0
    for row in range(2, ws.max_row + 1):
        gh = str(ws.cell(row=row, column=idx["GitHub"]).value or "").strip().lower().rstrip("/")
        conf = str(ws.cell(row=row, column=idx["GH置信度"]).value or "").lower()
        if conf == "medium" and gh in results:
            ws.cell(row=row, column=col, value=results[gh])
            n += 1
    wb.save(XLSX)
    print(f"回写 {n} 行（{len(results)} 人）到『GH复核』列")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--writeback", action="store_true")
    ap.add_argument("--workers", type=int, default=10)
    args = ap.parse_args()
    if args.writeback:
        writeback()
    else:
        run(args.workers)
