"""导出「传统 SDE · 未联系过」外联清单。

与 export_sde_list.py 的区别：
  1. 不限大厂白名单 —— 传统公司同样纳入
  2. 排除 Amazon / AWS 全系（含 Twitch、Audible 等子品牌）
  3. 排除 AI / ML / 研究岗（按职位词边界匹配，非按公司）
  4. 只要有 Email 的
  5. 只要没联系过的 —— people.status='new' 且无外联记录。
     外联记录的判定用三重兜底：person_id 命中、linkedin_url 命中、姓名命中。
     （history 里有 6 条 person_id 为空的历史脏数据，只查 person_id 会漏人）

用法:
    python3 ingest/export_sde_outreach_list.py          # 写 data/
"""
import sqlite3
import re
import datetime
from collections import Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DB = "data.db"

SDE = re.compile(
    r"\b(software\s+(engineer|developer|dev\b)|sde\b|software\s+development\s+engineer"
    r"|backend|back-end|frontend|front-end|full[\s-]?stack|programmer"
    r"|systems?\s+engineer|platform\s+engineer|infrastructure\s+engineer)", re.I)

# 排除 AI/研究岗。\b 词边界，避免 ai 命中 Email、ml 命中 HTML
AI_ROLE = re.compile(
    r"\b(ai|ml|a\.i\.|llm|nlp|genai|gen[\s-]?ai)\b"
    r"|machine\s+learning|deep\s+learning|research|scientist"
    r"|computer\s+vision|data\s+scien|applied\s+scien|perception|robotics", re.I)

AMAZON = re.compile(r"\b(amazon|aws|a9\.com|audible|twitch|zappos)\b", re.I)

# 学术职位。"CS & Software Engineering" 这类系名会被 SDE 规则误命中，需要挡掉。
# 不用裸 ph\.?d —— "Software Engineer, PhD" 是正经 SDE，只挡 PhD Student/Candidate
ACADEMIC = re.compile(r"教授|讲师|博士生|研究员|professor|lecturer|post[\s-]?doc"
                      r"|faculty|ph\.?\s?d\.?\s+(student|candidate)", re.I)

# 公司本身是 AI 研究机构 —— 职位是 SDE 但组织性质偏 AI，标出来让人工定夺
AI_ORG = re.compile(r"\b(deepmind|openai|anthropic|mistral|cohere|scale\s?ai"
                    r"|hugging\s?face|stability\s?ai|inflection)\b", re.I)


def norm_li(u):
    """LinkedIn URL 归一化，用于跨表比对。"""
    if not u:
        return ""
    u = re.sub(r"^https?://(www\.)?", "", u.strip().lower()).rstrip("/")
    return u


def main():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row

    # 已联系过的三重指纹
    contacted_ids, contacted_li, contacted_names = set(), set(), set()
    for h in conn.execute("SELECT person_id, name, url FROM history"):
        if h["person_id"]:
            contacted_ids.add(h["person_id"])
        if h["url"]:
            contacted_li.add(norm_li(h["url"]))
        if h["name"]:
            contacted_names.add(h["name"].strip().lower())

    rows = conn.execute("""
        SELECT id, first_name, last_name, title, company, location,
               email, linkedin_url, github_url, status
        FROM people
        WHERE email IS NOT NULL AND TRIM(email) != ''
    """).fetchall()

    funnel = Counter()
    kept = []
    for r in rows:
        funnel["1_有Email"] += 1
        title = r["title"] or ""
        if not SDE.search(title):
            continue
        funnel["2_SDE职位"] += 1
        if AI_ROLE.search(title):
            continue
        funnel["3_排除AI/研究岗"] += 1
        if ACADEMIC.search(title):
            continue
        funnel["3b_排除学术职位"] += 1
        if AMAZON.search(r["company"] or ""):
            continue
        funnel["4_排除Amazon/AWS"] += 1
        if (r["status"] or "new") != "new":
            continue
        funnel["5_status=new"] += 1
        name = f"{r['first_name'] or ''} {r['last_name'] or ''}".strip()
        if (r["id"] in contacted_ids
                or norm_li(r["linkedin_url"]) in contacted_li
                or name.lower() in contacted_names):
            continue
        funnel["6_无外联记录"] += 1
        kept.append({
            "姓名": name,
            "公司": r["company"] or "",
            "职位": title,
            "地点": r["location"] or "",
            "Email": r["email"],
            "LinkedIn": r["linkedin_url"] or "",
            "GitHub": r["github_url"] or "",
            "备注": "AI 研究机构，需人工确认" if AI_ORG.search(r["company"] or "") else "",
        })

    kept.sort(key=lambda x: (x["公司"], x["姓名"]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SDE未联系"
    headers = ["#", "姓名", "公司", "职位", "地点", "Email", "LinkedIn", "GitHub", "备注"]
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(vertical="center")

    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    for i, rec in enumerate(kept, 1):
        ws.append([i] + [rec[h] for h in headers[1:]])
        if rec["备注"]:
            for c in ws[ws.max_row]:
                c.fill = warn_fill

    for col, w in zip("ABCDEFGHI", [5, 20, 26, 34, 30, 34, 42, 34, 22]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    ts = datetime.date.today().isoformat()
    out = f"data/SDE未联系清单_{ts}.xlsx"
    wb.save(out)

    print("=== 漏斗 ===")
    for k in sorted(funnel):
        print(f"  {k:22} {funnel[k]:>6}")
    print(f"\nTOTAL={len(kept)}\nFILE={out}")
    print("\n公司 TOP 15:")
    for c, n in Counter(r["公司"] for r in kept).most_common(15):
        print(f"  {n:>5}  {c}")
    flagged = sum(1 for r in kept if r["备注"])
    print(f"\n标黄待确认(AI 研究机构): {flagged}")


if __name__ == "__main__":
    main()
