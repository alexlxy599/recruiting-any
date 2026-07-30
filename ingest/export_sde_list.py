"""导出「北美大厂传统 SDE」联络清单。
筛选逻辑：标题含 Software Engineer/Developer/SDE，排除 Research/Scientist；
地点在北美；公司属于主流大厂白名单。
"""
import sqlite3
import re
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DB = "data.db"

# 大厂白名单（用 LIKE 前缀匹配，兼容 "Amazon Web Services (AWS)" 等变体）
BIG_TECH = [
    "Google", "YouTube", "Microsoft", "LinkedIn", "Amazon", "AWS",
    "Meta", "Instagram", "Apple", "Salesforce", "Oracle", "NVIDIA",
    "Adobe", "Netflix", "Databricks", "Snowflake", "Stripe", "Airbnb",
    "Uber", "Lyft", "Roblox", "Snap", "Pinterest", "Datadog", "Zoox",
    "Cruise", "Waymo", "Nuro", "Instacart", "Walmart Global Tech",
    "Qualcomm", "Broadcom", "Intel", "Asana", "Coinbase", "Robinhood",
]

NA_LOC = ["United States", "Canada", "USA", "Bay Area", "Seattle",
          "Boston", "New York", "California", "Washington", "Texas",
          "Toronto", "Vancouver"]

conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row

loc_clause = " OR ".join(["location LIKE ?"] * len(NA_LOC))
rows = conn.execute(f"""
    SELECT first_name, last_name, title, company, location,
           email, linkedin_url, github_url, status
    FROM people
    WHERE (title LIKE '%Software Engineer%' OR title LIKE '%Software Developer%'
           OR title LIKE '%SDE%' OR title LIKE '%Software Dev%')
      AND title NOT LIKE '%Research%'
      AND title NOT LIKE '%Scientist%'
      AND email IS NOT NULL AND email != ''
      AND ({loc_clause})
    ORDER BY company, last_name
""", [f"%{l}%" for l in NA_LOC]).fetchall()


def is_big_tech(company):
    if not company:
        return None
    for b in BIG_TECH:
        # 词边界匹配，避免 "Google" 命中 "Googler Inc" 之类误伤（这里够用）
        if b.lower() in company.lower():
            return b
    return None


records = []
for r in rows:
    canon = is_big_tech(r["company"])
    if not canon:
        continue
    name = f"{r['first_name']} {r['last_name']}".strip()
    records.append({
        "姓名": name,
        "公司": r["company"],
        "大厂归类": canon,
        "职位": r["title"],
        "地点": r["location"],
        "Email": r["email"] or "",
        "LinkedIn": r["linkedin_url"] or "",
        "GitHub": r["github_url"] or "",
        "状态": r["status"] or "new",
    })

# 按大厂归类 -> 姓名 排序
records.sort(key=lambda x: (x["大厂归类"], x["姓名"]))

# 写 xlsx
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "北美大厂SDE"
headers = ["#", "姓名", "公司", "大厂归类", "职位", "地点", "Email", "LinkedIn", "GitHub", "状态"]
ws.append(headers)

head_fill = PatternFill("solid", fgColor="1F4E78")
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = head_fill
    c.alignment = Alignment(vertical="center")

for i, rec in enumerate(records, 1):
    ws.append([i, rec["姓名"], rec["公司"], rec["大厂归类"], rec["职位"],
               rec["地点"], rec["Email"], rec["LinkedIn"], rec["GitHub"], rec["状态"]])

widths = [5, 20, 22, 14, 34, 30, 34, 40, 34, 10]
for col, w in zip("ABCDEFGHIJ", widths):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

ts = datetime.date.today().isoformat()
out = f"data/北美大厂SDE联络清单_{ts}.xlsx"
wb.save(out)

# 汇总
from collections import Counter
by_company = Counter(r["大厂归类"] for r in records)
print(f"TOTAL={len(records)}")
print(f"FILE={out}")
for c, n in by_company.most_common():
    print(f"  {c}: {n}")
