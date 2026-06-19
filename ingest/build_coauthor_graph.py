"""从 ICML 总表生成共著关系边（零成本，纯本地）。

每篇论文 → 库内作者两两连边（可点击）+ 连上全部共著者名字（待对齐）。
匹配人靠 GitHub URL（不靠名字，避开同名/一GitHub挂多人的坑）。
写入 collaborations，source='icml_coauthor'，relation='coauthor'。
"""

import os
import re
import sys
from collections import defaultdict

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import db

XLSX = "/Users/alex/Desktop/ICML 2026 分工版.xlsx"


def parse_names(detail: str) -> list[str]:
    """'Name [Org|label], Name [..], ..' → [Name, ...]"""
    out = []
    for part in re.split(r"\],\s*", (detail or "").strip()):
        part = part.rstrip("]").strip()
        if not part:
            continue
        m = re.match(r"^(.*?)\s*\[", part)
        out.append((m.group(1) if m else part).strip())
    return [n for n in out if n]


def main():
    conn = db.get_conn()
    # github_url(归一) → person_id
    gh2pid = {}
    for r in conn.execute("SELECT id, github_url FROM people WHERE github_url!=''"):
        gh2pid[r["github_url"].strip().lower().rstrip("/")] = r["id"]

    wb = openpyxl.load_workbook(XLSX, read_only=True)

    # 1. 每篇论文里：库内作者 {name_lower: pid}（靠 GitHub 匹配）
    ws_a = wb["全部华人作者"]
    rows = ws_a.iter_rows(values_only=True)
    idx = {h: i for i, h in enumerate(next(rows))}
    paper_db = defaultdict(dict)   # rank -> {author_lower: pid}
    for r in rows:
        rank = str(r[idx["Rank"]] or "").strip()
        author = str(r[idx["Author"]] or "").strip()
        gh = str(r[idx["GitHub"]] or "").strip().lower().rstrip("/")
        pid = gh2pid.get(gh)
        if rank and author and pid:
            paper_db[rank][author.lower()] = pid

    # 2. 每篇论文全作者名单
    ws = wb["文章名单"]
    rows = ws.iter_rows(values_only=True)
    hidx = {h: i for i, h in enumerate(next(rows))}
    paper_authors = {}
    titles = {}
    for r in rows:
        rank = str(r[hidx["Rank"]] or "").strip()
        if rank:
            paper_authors[rank] = parse_names(r[hidx["Authors Detail"]])
            titles[rank] = (r[hidx["Title"]] or "")[:150]

    # 3. 生成边：每个库内作者 → 同篇其余作者
    conn.execute("DELETE FROM collaborations WHERE source='icml_coauthor'")
    edges = defaultdict(dict)   # pid -> {coauthor_name: (co_pid, context)}
    for rank, db_authors in paper_db.items():
        names = paper_authors.get(rank, [])
        name_to_pid = {n.lower(): db_authors.get(n.lower()) for n in names}
        for author_lower, pid in db_authors.items():
            for co in names[:30]:
                if co.lower() == author_lower:
                    continue
                co_pid = name_to_pid.get(co.lower())
                if co.lower() not in edges[pid]:
                    edges[pid][co.lower()] = (co, co_pid, titles.get(rank, ""))

    n = 0
    for pid, cos in edges.items():
        for _, (name, co_pid, ctx) in cos.items():
            conn.execute(
                """INSERT INTO collaborations (person_id, collaborator_name, collaborator_person_id,
                                               relation, context, source)
                   VALUES (?,?,?, 'coauthor', ?, 'icml_coauthor')""",
                (pid, name[:80], co_pid, f"ICML 共著: {ctx}"))
            n += 1
    conn.commit()

    linked = conn.execute(
        "SELECT COUNT(*) FROM collaborations WHERE source='icml_coauthor' AND collaborator_person_id IS NOT NULL"
    ).fetchone()[0]
    ppl = conn.execute(
        "SELECT COUNT(DISTINCT person_id) FROM collaborations WHERE source='icml_coauthor'").fetchone()[0]
    conn.close()
    print(f"生成共著边 {n} 条，覆盖 {ppl} 人，其中 {linked} 条连到库内可点击节点")


if __name__ == "__main__":
    main()
