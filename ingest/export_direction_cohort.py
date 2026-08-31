"""按研究方向导出候选人清单,按「管道语义 × 顶会信号」分层。

与早期版本的关键差异:管道状态不看 history.status(几乎全是 'replied'),
看 history.reply 的**实际语义**。实测 290 条回复里,"沟通中"和"暂无兴趣"
都记作 replied —— 混在一起排序会把明确拒绝的人排到名单最前面。

用法:
    python3.12 ingest/export_direction_cohort.py                    # 默认多模态/AIGC
    python3.12 ingest/export_direction_cohort.py --name 具身智能 --kw embodied,robotics
"""
import argparse
import datetime
import re
import sys
import os
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import db

# ── 方向关键词 ──
DEFAULT_STRONG = [
    "multimodal", "multi-modal", "vision-language", "vision language", "vlm", "mllm",
    "aigc", "text-to-image", "text to image", "image generation", "video generation",
    "image synthesis", "video synthesis", "diffusion",
    "多模态", "图像生成", "视频生成", "文生图", "文生视频",
]
DEFAULT_WEAK = ["generative", "gan ", "stable diffusion", "生成式"]

# ── 回复语义分类 ──
# 顺序即优先级:先命中的先判定。拒绝类必须排在"保持联系"之前,
# 否则"暂无兴趣,保持联系"会被误判成正面。
REPLY_RULES = [
    ("拒绝·地域", r"只对美国|只考虑美国|只对加拿大|暂时没有回国|暂不回国|短期还不想回国"),
    ("拒绝·明确", r"暂无兴趣|不感兴趣|暂无匹配岗|不合适|难以匹配|非目标|暂无计划|"
                  r"短期不找工作|暂时不看|不来|暂无意愿|暂无回复|怪人"),
    ("流程中断", r"半路退出流程|部门流转中|未出席|临时不来|没约上|不回邮件"),
    ("推进·简历", r"等待简历|答应给简历|给了简历|有简历|待推荐|待综面|简历更新|待业务交流|业务面试中"),
    ("推进·约谈", r"沟通中|约时间|拟定周|现场见面聊|可以聊聊|愿意聊|线上|电话|"
                  r"随时约|再约|约见面|有意愿|感兴趣|主动"),
    ("时间敏感", r"正在面试|再试一年教职|找教职|在.*流程中"),
    ("弱正·留门", r"保持联系|微信已加|加了微信|加微信|ICML见|nips|NIPS|后联系|再说|"
                  r"过些时间联系|明年|九月|等女友"),
]
# 分层优先级(数字小=优先)
TIER_ORDER = {
    "S·推进中": 0, "S·时间敏感": 1, "A·留门可再触": 2,
    "A·已触达待回": 3,
    "B·顶会一作(未联系)": 4, "C·顶会相关(未联系)": 5,
    "D·仅方向(未联系)": 6, "X·流程中断": 7, "Z·已拒绝": 8,
}
TIER_FILL = {
    "S·推进中": "C6EFCE", "S·时间敏感": "FFE699", "A·留门可再触": "E2EFDA",
    "A·已触达待回": "DDEBF7", "B·顶会一作(未联系)": "FFF2CC",
    "X·流程中断": "F2F2F2", "Z·已拒绝": "FCE4E4",
}


def classify_reply(text):
    t = (text or "").strip()
    if not t:
        return None
    for label, pat in REPLY_RULES:
        if re.search(pat, t, re.I):
            return label
    return "其他·需人读"


def collect(conn, strong, weak):
    """方向命中 → {pid: {"score": int, "kw": set}}"""
    hits = {}
    def bump(pid, w, kw):
        d = hits.setdefault(pid, {"score": 0, "kw": set()})
        d["score"] += w
        d["kw"].add(kw.strip())

    for kws, w in ((strong, 2), (weak, 1)):
        for kw in kws:
            like = f"%{kw}%"
            for (pid,) in conn.execute(
                "SELECT DISTINCT person_id FROM person_tags pt JOIN tags t ON t.id=pt.tag_id "
                "WHERE LOWER(t.name) LIKE ?", (like,)):
                bump(pid, w, kw)
            for (pid,) in conn.execute(
                "SELECT id FROM people WHERE LOWER(COALESCE(research_area,'')) LIKE ?", (like,)):
                bump(pid, w, kw)
            for (pid,) in conn.execute(
                "SELECT id FROM people WHERE LOWER(COALESCE(headline,'')||' '||COALESCE(title,'')) LIKE ?",
                (like,)):
                bump(pid, w, kw)
            for (pid,) in conn.execute(
                "SELECT DISTINCT person_id FROM publications WHERE LOWER(title) LIKE ? "
                "AND person_id IS NOT NULL", (like,)):
                bump(pid, w, kw)
    return hits


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--name", default="多模态AIGC")
    ap.add_argument("--kw", help="逗号分隔,覆盖默认强关键词")
    ap.add_argument("--limit-cold", type=int, default=120, help="未联系过的人保留上限")
    a = ap.parse_args()

    strong = [x.strip() for x in a.kw.split(",")] if a.kw else DEFAULT_STRONG
    weak = [] if a.kw else DEFAULT_WEAK

    conn = db.get_conn()
    hits = collect(conn, strong, weak)
    ids = sorted(hits)
    ph = ",".join("?" * len(ids))

    info = {r["id"]: r for r in conn.execute(
        f"""SELECT id, first_name, last_name, company, title, institution, headline,
                   email, linkedin_url, sector, status
            FROM people WHERE id IN ({ph})""", ids)}

    pubs = defaultdict(list)
    for r in conn.execute(
        f"""SELECT person_id, venue, year, is_first_author, title
            FROM publications WHERE person_id IN ({ph}) AND COALESCE(venue,'')!=''""", ids):
        pubs[r["person_id"]].append(r)

    # 管道:取每人最新一条有内容的回复
    latest_reply = {}
    for r in conn.execute(
        f"""SELECT person_id, reply, created_at FROM history
            WHERE person_id IN ({ph}) AND reply IS NOT NULL AND TRIM(reply)!=''
            ORDER BY person_id, created_at""", ids):
        latest_reply[r["person_id"]] = r
    touch_count = {r["person_id"]: r["n"] for r in conn.execute(
        f"SELECT person_id, COUNT(*) n FROM history WHERE person_id IN ({ph}) GROUP BY person_id", ids)}

    rows = []
    for pid in ids:
        p = info.get(pid)
        if not p:
            continue
        d = hits[pid]
        pu = pubs.get(pid, [])
        fa = [x for x in pu if x["is_first_author"]]
        vc = Counter(x["venue"] for x in pu)
        venues = " ".join(f"{v}×{c}" if c > 1 else v for v, c in vc.most_common(4))

        rp = latest_reply.get(pid)
        sem = classify_reply(rp["reply"]) if rp else None

        if sem in ("推进·简历", "推进·约谈"):
            tier = "S·推进中"
        elif sem == "时间敏感":
            tier = "S·时间敏感"
        elif sem in ("弱正·留门", "其他·需人读"):
            tier = "A·留门可再触"
        elif sem == "流程中断":
            tier = "X·流程中断"
        elif sem and sem.startswith("拒绝"):
            tier = "Z·已拒绝"
        elif touch_count.get(pid, 0) > 0:
            # 发过但没记回复 —— 既不是冷线索也不是热线索,要人工看是否跟进
            tier = "A·已触达待回"
        elif fa:
            tier = "B·顶会一作(未联系)"
        elif pu:
            tier = "C·顶会相关(未联系)"
        else:
            tier = "D·仅方向(未联系)"

        rows.append({
            "tier": tier, "order": TIER_ORDER[tier],
            "score": d["score"] + 3 * len(fa) + len(pu),
            "pid": pid,
            "姓名": f"{p['first_name'] or ''} {p['last_name'] or ''}".strip(),
            "公司/机构": p["company"] or p["institution"] or "",
            "职位": p["title"] or "",
            "方向信号": ", ".join(sorted(d["kw"])[:5]),
            "顶会": venues, "一作": len(fa),
            "代表作": ((fa[0]["title"] if fa else (pu[0]["title"] if pu else "")) or "")[:70],
            "上次回复": (rp["reply"][:60] if rp else ""),
            "回复判定": sem or "",
            "触达次数": touch_count.get(pid, 0),
            "Email": p["email"] or "", "LinkedIn": p["linkedin_url"] or "",
        })

    rows.sort(key=lambda x: (x["order"], -x["score"]))
    warm = [r for r in rows if r["order"] <= 3]
    cold = [r for r in rows if 4 <= r["order"] <= 6][:a.limit_cold]
    dead = [r for r in rows if r["order"] >= 7]
    keep = warm + cold

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = a.name[:28]
    headers = ["#", "分层", "姓名", "公司/机构", "职位", "方向信号", "顶会", "一作",
               "代表作", "上次回复", "回复判定", "触达次数", "Email", "LinkedIn", "person_id"]

    def write(sheet, data):
        sheet.append(headers)
        hf = PatternFill("solid", fgColor="1F4E78")
        for c in sheet[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = hf
            c.alignment = Alignment(vertical="center")
        for i, r in enumerate(data, 1):
            sheet.append([i, r["tier"], r["姓名"], r["公司/机构"], r["职位"], r["方向信号"],
                          r["顶会"], r["一作"], r["代表作"], r["上次回复"], r["回复判定"],
                          r["触达次数"], r["Email"], r["LinkedIn"], r["pid"]])
            f = TIER_FILL.get(r["tier"])
            if f:
                for c in sheet[sheet.max_row]:
                    c.fill = PatternFill("solid", fgColor=f)
        for col, w in zip("ABCDEFGHIJKLMNO",
                          [5, 19, 18, 28, 24, 30, 24, 6, 52, 34, 13, 9, 28, 38, 10]):
            sheet.column_dimensions[col].width = w
        sheet.freeze_panes = "C2"

    write(ws, keep)
    write(wb.create_sheet("勿再触达"), dead)

    ts = datetime.date.today().isoformat()
    os.makedirs("data", exist_ok=True)
    out = f"data/{a.name}候选人_{ts}.xlsx"
    wb.save(out)

    print(f"方向命中 {len(rows)} 人 → 导出 {len(keep)}(主表) + {len(dead)}(勿再触达)")
    print(f"FILE={out}\n")
    print("── 分层 ──")
    for t, n in sorted(Counter(r["tier"] for r in rows).items(), key=lambda x: TIER_ORDER[x[0]]):
        mark = "  ← 主表" if TIER_ORDER[t] <= 6 else "  ← 独立sheet"
        print(f"  {t:<22} {n:>4}{mark}")
    print("\n── 回复语义分布(管道内) ──")
    for s, n in Counter(r["回复判定"] for r in rows if r["回复判定"]).most_common():
        print(f"  {s:<12} {n}")
    print("\n── S/A 层全量(优先联系) ──")
    for r in warm:
        print(f"  {r['tier'][:1]} #{r['pid']:<5} {r['姓名'][:15]:17s} {(r['公司/机构'] or '—')[:22]:24s} "
              f"{r['顶会'][:16]:18s} {r['上次回复'][:26]}")


if __name__ == "__main__":
    main()
