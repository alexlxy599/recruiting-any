"""为名单里的候选人抓「最新动向」:个人主页 News 区 + GitHub 活跃度。

三个数据源,可得性差别很大:
  个人主页  —— 直接抓,学术主页的 News/Updates 区是最好的动向来源
  GitHub   —— 需要有效 token(无认证只有 60 次/小时)
  LinkedIn —— HTTP 999 反爬拦截,**不绕过**,一律标记为未更新

抓取结果过 enrichment_cache(30天),重跑不会重复打外部接口。

用法:
    python3.12 ingest/enrich_status.py --pages          # 只抓主页
    python3.12 ingest/enrich_status.py --github         # 只抓 GitHub
    python3.12 ingest/enrich_status.py --pages --github # 全跑
"""
import argparse
import json
import os
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from dotenv import load_dotenv

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import db

load_dotenv()
UA = {"User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                     "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36")}
LLM_URL = os.environ.get("OPENAI_BASE_URL", "http://127.0.0.1:1234/v1")
LLM_MODEL = os.environ.get("DEFAULT_MODEL", "qwen3.5-9b-deepseek-v4-flash")
OUT = "data/raw/status_enrich.json"

NEWS_HINT = re.compile(r"\b(news|updates?|recent|what'?s new|近况|动态|最新)\b", re.I)


# ── 个人主页 ──
def fetch_page(url):
    r = requests.get(url, headers=UA, timeout=20, allow_redirects=True)
    if r.status_code != 200:
        return None
    html = re.sub(r"<(script|style|nav|footer)[^>]*>.*?</\1>", " ", r.text, flags=re.S | re.I)
    text = re.sub(r"<[^>]+>", "\n", html)
    text = re.sub(r"&[a-z#0-9]+;", " ", text)
    lines = [re.sub(r"\s+", " ", l).strip() for l in text.split("\n")]
    lines = [l for l in lines if len(l) > 2]
    # News 区优先:命中标题后往下取 40 行
    for i, l in enumerate(lines):
        if NEWS_HINT.search(l) and len(l) < 40:
            return "\n".join(lines[i:i + 40])[:2600]
    return "\n".join(lines[:60])[:2600]


PROMPT = """你在读一位研究者的个人主页,为招聘做背景更新。只依据给定文本,不要编造。

输出严格 JSON,字段:
  current: 当前身份与所在机构(如 "PhD student @ MIT" / "Research Scientist @ Google"),没有写 ""
  latest: 最近的动向,一句话(如 "2026年6月起在NVIDIA实习" / "2篇论文被NeurIPS 2026接收"),没有写 ""
  grad: 预计毕业时间,只在文本明确提到时填(如 "2027"),否则 ""
  seeking: 是否提到在找工作/实习/教职,是则简述,否则 ""

主页文本:
---
{text}
---
只输出 JSON,不要解释。"""


def llm_extract(text):
    try:
        r = requests.post(f"{LLM_URL}/chat/completions", timeout=90, json={
            "model": LLM_MODEL, "temperature": 0,
            "messages": [{"role": "user", "content": PROMPT.format(text=text)}],
        })
        if r.status_code != 200:
            return None
        c = r.json()["choices"][0]["message"]["content"]
        m = re.search(r"\{.*\}", c, re.S)
        return json.loads(m.group()) if m else None
    except Exception:
        return None


def do_pages(targets, workers=6):
    out, done = {}, 0
    def work(t):
        pid, url = t
        cached = db.cache_get(f"status_page:{url}")
        if cached:
            return pid, json.loads(cached), "cache"
        txt = fetch_page(url)
        if not txt:
            return pid, None, "fetch_fail"
        data = llm_extract(txt)
        if data:
            db.cache_set(f"status_page:{url}", json.dumps(data, ensure_ascii=False))
        return pid, data, "ok" if data else "llm_fail"

    with ThreadPoolExecutor(max_workers=workers) as ex:
        futs = {ex.submit(work, t): t for t in targets}
        for f in as_completed(futs):
            pid, data, how = f.result()
            done += 1
            if data:
                out[pid] = data
            if done % 10 == 0 or done == len(targets):
                print(f"  主页 {done}/{len(targets)}  成功 {len(out)}", flush=True)
    return out


# ── GitHub ──
def gh_headers():
    tok = (os.environ.get("GITHUB_TOKEN") or "").strip()
    h = dict(UA)
    h["Accept"] = "application/vnd.github+json"
    if tok:
        h["Authorization"] = f"Bearer {tok}"
    return h, bool(tok)


def gh_login(url):
    m = re.search(r"github\.com/([A-Za-z0-9\-_.]+)", url or "")
    if not m:
        return None
    lg = m.group(1)
    return None if lg.lower() in ("orgs", "settings", "about") else lg


def do_github(targets):
    h, has_tok = gh_headers()
    r = requests.get("https://api.github.com/rate_limit", headers=h, timeout=15)
    if r.status_code == 401:
        print("!! GITHUB_TOKEN 无效(401)。改用无认证模式,限额 60次/小时")
        h.pop("Authorization", None)
        has_tok = False
        r = requests.get("https://api.github.com/rate_limit", headers=h, timeout=15)
    core = r.json().get("resources", {}).get("core", {})
    remaining = core.get("remaining", 0)
    print(f"  GitHub 额度: {remaining}/{core.get('limit')}  ({'已认证' if has_tok else '无认证'})")
    if remaining < len(targets):
        print(f"  !! 额度不足({remaining} < {len(targets)}),本轮只处理前 {remaining} 人,其余下轮续跑")
        targets = targets[:max(0, remaining - 2)]

    out = {}
    for i, (pid, url) in enumerate(targets, 1):
        lg = gh_login(url)
        if not lg:
            continue
        ck = f"status_gh:{lg}"
        cached = db.cache_get(ck)
        if cached:
            out[pid] = json.loads(cached)
            continue
        try:
            u = requests.get(f"https://api.github.com/users/{lg}", headers=h, timeout=20)
            if u.status_code != 200:
                continue
            j = u.json()
            ev = requests.get(f"https://api.github.com/users/{lg}/events/public?per_page=30",
                              headers=h, timeout=20)
            evs = ev.json() if ev.status_code == 200 else []
            repos = []
            for e in evs:
                if e.get("type") in ("PushEvent", "CreateEvent", "PullRequestEvent"):
                    rn = (e.get("repo") or {}).get("name", "")
                    if rn and rn not in repos:
                        repos.append(rn)
            data = {
                "gh_name": j.get("name") or "",
                "gh_company": j.get("company") or "",
                "gh_bio": j.get("bio") or "",
                "gh_location": j.get("location") or "",
                "gh_blog": j.get("blog") or "",
                "gh_updated": (j.get("updated_at") or "")[:10],
                "gh_last_active": (evs[0].get("created_at") or "")[:10] if evs else "",
                "gh_recent_repos": ", ".join(repos[:3]),
                "gh_followers": j.get("followers", 0),
            }
            db.cache_set(ck, json.dumps(data, ensure_ascii=False))
            out[pid] = data
        except Exception:
            pass
        if not has_tok:
            time.sleep(1.2)
        if i % 10 == 0:
            print(f"  GitHub {i}/{len(targets)}  成功 {len(out)}", flush=True)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pages", action="store_true")
    ap.add_argument("--github", action="store_true")
    ap.add_argument("--xlsx", default="data/多模态AIGC名单_2026-08-26.xlsx")
    a = ap.parse_args()
    if not (a.pages or a.github):
        ap.print_help(); return

    import openpyxl
    ws = openpyxl.load_workbook(a.xlsx).active
    H = [c.value for c in ws[1]]
    rows = [dict(zip(H, [c.value for c in r])) for r in ws.iter_rows(min_row=2)]

    prev = json.load(open(OUT, encoding="utf-8")) if os.path.exists(OUT) else {}
    merged = {str(k): v for k, v in prev.items()}

    if a.pages:
        t = [(r["id"], r["个人主页"]) for r in rows
             if r.get("个人主页") and str(r["个人主页"]).startswith("http")]
        print(f"主页待抓 {len(t)} 人")
        for pid, d in do_pages(t).items():
            merged.setdefault(str(pid), {}).update(d)

    if a.github:
        t = [(r["id"], r["GitHub"]) for r in rows
             if r.get("GitHub") and str(r["GitHub"]).startswith("http")]
        print(f"GitHub 待抓 {len(t)} 人")
        for pid, d in do_github(t).items():
            merged.setdefault(str(pid), {}).update(d)

    os.makedirs("data/raw", exist_ok=True)
    json.dump(merged, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print(f"\n累计 {len(merged)} 人有动向数据 → {OUT}")


if __name__ == "__main__":
    main()
