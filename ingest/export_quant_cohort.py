"""导出「ICML 2026 量化方向」候选人清单。

两个来源合并去重：
  1. 库内已有 —— publications.title / people.research_area 命中严格量化关键词
  2. 本次新增 —— 从 icml.cc/virtual 抓的 ICML 2026 量化论文作者，过华人姓氏筛后
     不在库内的人（还没入库，person_id 为空）

分档口径（可解释，不用 LLM 打分）：
  B    一作          —— 优先联系
  C    中间作者      —— 入库养着
  导师  末位作者      —— 不当候选人，当实验室入口

用法:
    python3.12 ingest/export_quant_cohort.py
"""
import datetime
import json
import os
import re
import sqlite3
import sys
from collections import Counter

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from conference_scraper import has_chinese_surname

DB = "data.db"
COHORT_JSON = os.environ.get("QUANT_COHORT_JSON", "data/raw/icml2026_quant_cohort.json")

# 严格量化口径。不含 pruning / distillation / KV cache —— 那是压缩大类，方向会散
STRICT_KW = [
    "quantiz", "low-bit", "low bit", "int8", "int4", "fp8", "bitwidth", "bit-width",
    "post-training quant", "gptq", "awq", "smoothquant", "binariz", "ternary",
    "1-bit", "bitnet", "mixed-precision", "mixed precision",
]


def norm_name(s):
    return re.sub(r"[^a-z]", "", (s or "").lower())


def grade_of(is_first, is_last):
    if is_last and not is_first:
        return "导师"
    return "B" if is_first else "C"


def from_db(conn):
    """库内已有的量化 cohort。"""
    ids = set()
    for kw in STRICT_KW:
        like = f"%{kw}%"
        ids |= {r[0] for r in conn.execute(
            "SELECT DISTINCT person_id FROM publications WHERE LOWER(title) LIKE ?", (like,))}
        ids |= {r[0] for r in conn.execute(
            "SELECT id FROM people WHERE LOWER(COALESCE(research_area,'')) LIKE ?", (like,))}
    if not ids:
        return []
    ph = ",".join("?" * len(ids))
    rows = conn.execute(f"""
        SELECT id, first_name, last_name, email, institution, personal_page,
               title, status, github_verified, sector
        FROM people WHERE id IN ({ph})""", sorted(ids)).fetchall()

    # 代表论文必须是"让他入选量化 cohort 的那篇"，否则表上会显示一篇无关论文，
    # 看的人无从判断这人为什么在名单里。命中量化关键词的优先，其次才退回一作论文。
    kw_sql = " OR ".join(["LOWER(title) LIKE ?"] * len(STRICT_KW))
    kw_params = [f"%{k}%" for k in STRICT_KW]

    out = []
    for r in rows:
        pub = conn.execute(f"""
            SELECT title, venue, year, is_first_author FROM publications
            WHERE person_id = ? AND ({kw_sql})
            ORDER BY is_first_author DESC, year DESC LIMIT 1""",
            [r["id"]] + kw_params).fetchone()
        matched_by = "论文" if pub else ""
        if not pub:
            # 靠 research_area 入选的，论文里没有量化题目 —— 标出来，别混为一谈
            pub = conn.execute("""
                SELECT title, venue, year, is_first_author FROM publications
                WHERE person_id = ? ORDER BY is_first_author DESC, year DESC LIMIT 1""",
                (r["id"],)).fetchone()
            matched_by = "研究方向"
        out.append({
            "person_id": r["id"],
            "姓名": f"{r['first_name'] or ''} {r['last_name'] or ''}".strip(),
            "分档": grade_of(bool(pub and pub["is_first_author"]), False),
            "机构": r["institution"] or "",
            "Email": r["email"] or "",
            "个人主页": r["personal_page"] or "",
            "代表论文": (pub["title"] if pub else "") or "",
            "会议": f"{pub['venue']} {pub['year']}" if pub and pub["venue"] else "",
            "作者位": "一作" if pub and pub["is_first_author"] else "",
            "状态": r["status"] or "new",
            "来源": "库内已有",
            "论文链接": "",
            "备注": "" if matched_by == "论文" else "按研究方向入选，论文题目无量化字样",
        })
    return out


ATTR = "data/raw/icml2026_quant_attribution.json"


def load_attribution():
    """arXiv 归属结果：姓名 → 机构/邮箱。键用 squash 过的名字，容忍写法差异。"""
    if not os.path.exists(ATTR):
        return {}
    raw = json.load(open(ATTR, encoding="utf-8")).get("attribution", {})
    return {norm_name(k): {**v, "_name": k} for k, v in raw.items()}


def from_web(path):
    """本次从 icml.cc 抓的新人。"""
    if not os.path.exists(path):
        print(f"!! 找不到 {path}，跳过新增部分")
        return []
    papers = json.load(open(path, encoding="utf-8"))
    attr = load_attribution()

    people = {}
    for title, poster, tier, authors in papers:
        n = len(authors)
        for i, a in enumerate(authors):
            d = people.setdefault(a.lower().strip(), {"name": a, "papers": []})
            d["papers"].append({"title": title, "tier": tier, "poster": poster,
                                "pos": i, "nauth": n})

    out = []
    for d in people.values():
        if not has_chinese_surname(d["name"]):
            continue
        first = any(p["pos"] == 0 for p in d["papers"])
        last = any(p["pos"] == p["nauth"] - 1 and p["nauth"] > 2 for p in d["papers"])
        p = sorted(d["papers"], key=lambda x: x["pos"])[0]
        pos_label = "一作" if p["pos"] == 0 else (
            "末位" if p["pos"] == p["nauth"] - 1 else f"第{p['pos']+1}作者")
        a = attr.get(norm_name(d["name"])) or {}
        # arXiv 上的写法更规范（icml.cc 有 HuangJunTao / Lilaiyi 这类黏连写法）
        display = a.get("_name") or d["name"]
        note = a.get("corrected", "")
        if a.get("confidence") == "low":
            note = (note + " ｜ 低置信").strip(" ｜")
        out.append({
            "person_id": "",
            "姓名": display,
            "分档": grade_of(first, last) if not a.get("is_last") else "导师",
            "机构": a.get("institution", ""),
            "Email": a.get("email", ""),
            "个人主页": "",
            "代表论文": p["title"],
            "会议": f"ICML 2026 {p['tier']}",
            "作者位": pos_label,
            "状态": "未入库",
            "来源": "本次新增",
            "论文链接": f"https://icml.cc/virtual/2026/poster/{p['poster']}",
            "备注": note,
        })
    return out


def main():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row

    db_rows = from_db(conn)
    web_rows = from_web(COHORT_JSON)

    # 全库姓名索引。只跟严格量化那批比是不够的 —— web 抓到的人可能早就在库里，
    # 只是论文标题没命中量化关键词。漏掉这层比对，导入时会造出重复的 people 行。
    all_people = {}
    for r in conn.execute("SELECT id, first_name, last_name, institution, email, "
                          "personal_page, status FROM people"):
        k = norm_name(f"{r['first_name'] or ''}{r['last_name'] or ''}")
        if k:
            all_people.setdefault(k, r)

    seen = {norm_name(r["姓名"]) for r in db_rows}
    merged = list(db_rows)
    dropped = 0
    rescued = 0
    for r in web_rows:
        k = norm_name(r["姓名"])
        if k in seen:
            dropped += 1
            continue
        seen.add(k)
        hit = all_people.get(k)
        if hit:
            # 库里已有此人，补上 person_id 与已知字段，别当新人导
            r["person_id"] = hit["id"]
            r["机构"] = hit["institution"] or ""
            r["Email"] = hit["email"] or ""
            r["个人主页"] = hit["personal_page"] or ""
            r["状态"] = hit["status"] or "new"
            r["来源"] = "库内已有(量化关键词未命中)"
            rescued += 1
        merged.append(r)

    # 库内存量邮箱质量参差：有掩码残留(****@)、有 QQ 号邮箱、有域名与机构对不上的。
    # 不删 —— 可能是本人的私人/母校邮箱 —— 但必须标出来，避免当成可直接外联的地址。
    for r in merged:
        e = (r.get("Email") or "").strip()
        if not e:
            continue
        flag = ""
        if "*" in e:
            flag = "邮箱被掩码，不可用"
        elif re.match(r"^\d{6,}@qq\.com$", e, re.I):
            flag = "QQ 号邮箱，可用性存疑"
        else:
            inst = (r.get("机构") or "").lower()
            dom = e.split("@")[-1].lower()
            hint = re.split(r"[.@]", dom)[0]
            if inst and len(hint) > 3 and hint not in ("gmail", "outlook", "163", "qq", "foxmail") \
               and hint not in inst.replace(" ", "") and hint not in inst:
                flag = f"邮箱域名({dom})与机构不符，需核对"
        if flag:
            r["备注"] = (r.get("备注") or "")
            r["备注"] = (r["备注"] + " ｜ " + flag).strip(" ｜")

    order = {"B": 0, "C": 1, "导师": 2}
    merged.sort(key=lambda x: (order.get(x["分档"], 9), x["来源"] != "库内已有", x["姓名"]))

    headers = ["#", "姓名", "分档", "机构", "Email", "个人主页",
               "代表论文", "会议", "作者位", "状态", "来源", "person_id", "论文链接", "备注"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "量化cohort"
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
        c.alignment = Alignment(vertical="center")

    b_fill = PatternFill("solid", fgColor="E2EFDA")      # 一作，绿
    mentor_fill = PatternFill("solid", fgColor="FCE4D6")  # 导师，橙
    todo_fill = PatternFill("solid", fgColor="FFF2CC")    # 待 enrich，黄

    for i, rec in enumerate(merged, 1):
        ws.append([i] + [rec.get(h, "") for h in headers[1:]])
        row = ws[ws.max_row]
        if rec["分档"] == "B":
            row[2].fill = b_fill
        elif rec["分档"] == "导师":
            row[2].fill = mentor_fill
        if not rec["机构"] and not rec["Email"]:
            row[3].fill = todo_fill
            row[4].fill = todo_fill

    for col, w in zip("ABCDEFGHIJKLMN",
                      [5, 20, 7, 34, 30, 34, 58, 20, 11, 9, 22, 11, 44, 46]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "C2"

    ts = datetime.date.today().isoformat()
    os.makedirs("data", exist_ok=True)
    out = f"data/ICML2026量化cohort_{ts}.xlsx"
    wb.save(out)

    print("=== 来源 ===")
    print(f"  库内已有            {len(db_rows):>5}")
    print(f"  web 抓取(华人筛后)  {len(web_rows):>5}")
    print(f"  姓名撞车丢弃        {dropped:>5}")
    print(f"  web命中库内已有人   {rescued:>5}  (已回填 person_id，勿重复导入)")
    print(f"  合并去重后          {len(merged):>5}")
    todo = sum(1 for r in merged if not r["person_id"])
    print(f"  其中待入库          {todo:>5}")
    print("\n=== 分档 ===")
    for g, n in sorted(Counter(r["分档"] for r in merged).items(),
                       key=lambda x: order.get(x[0], 9)):
        print(f"  {g:<5} {n:>5}")
    print("\n=== 数据完整度 ===")
    for label, key in [("有机构", "机构"), ("有Email", "Email"), ("有主页", "个人主页")]:
        n = sum(1 for r in merged if r[key])
        print(f"  {label:<8} {n:>5} / {len(merged)}  ({n*100//len(merged)}%)")
    print(f"\nFILE={out}")
    print("\n机构 TOP 12:")
    for c, n in Counter(r["机构"] for r in merged if r["机构"]).most_common(12):
        print(f"  {n:>4}  {c}")


if __name__ == "__main__":
    main()
