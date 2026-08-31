"""人工(Opus)读主页快照 → 结构化提取的读写通道。

分工:
    dump    取一批待提取的人,输出快照正文供阅读
    load    把阅读得出的 JSON 写进 extractions 并跑 normalize() 落规范层

为什么不用批量 LLM:提取质量直接决定人才库画像的可信度,而主页格式极杂
(导师/合作者/项目/毕业年散落在各处,还常有模板残留文本)。这条通道让判断
由人做,脚本只负责搬运。

用法:
    python3.12 ingest/manual_extract.py dump --n 30 [--cohort 文件.xlsx]
    python3.12 ingest/manual_extract.py load  batch.json
"""
import argparse
import json
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import db
from enrich_homepage import EXTRACT_VERSION, VERIFIED, normalize

MODEL_TAG = "opus-manual"


def pending(conn, n, priority_ids=None):
    ph = ",".join("?" * len(VERIFIED))
    sql = f"""
        SELECT p.id, p.first_name || ' ' || p.last_name AS name,
               COALESCE(p.institution, p.company, '') AS aff, p.personal_page
        FROM people p
        WHERE p.github_verified IN ({ph})
          AND p.id IN (SELECT DISTINCT person_id FROM web_snapshots WHERE source='homepage')
          AND p.id NOT IN (SELECT person_id FROM extractions
                           WHERE source='homepage' AND version=?)"""
    params = list(VERIFIED) + [EXTRACT_VERSION]
    if priority_ids:
        pid_ph = ",".join("?" * len(priority_ids))
        sql += f" ORDER BY CASE WHEN p.id IN ({pid_ph}) THEN 0 ELSE 1 END, p.id"
        params += list(priority_ids)
    else:
        sql += " ORDER BY p.id"
    sql += " LIMIT ?"
    params.append(n)
    return [dict(r) for r in conn.execute(sql, params)]


def snapshot_text(conn, pid, cap=1100):
    r = conn.execute("""SELECT raw_text AS content FROM web_snapshots
                        WHERE person_id=? AND source='homepage'
                        ORDER BY id DESC LIMIT 1""", (pid,)).fetchone()
    if not r or not r["content"]:
        return ""
    t = r["content"]
    if "<" in t[:400]:
        t = re.sub(r"<(script|style|nav|footer|head)[^>]*>.*?</\1>", " ", t, flags=re.S | re.I)
        t = re.sub(r"<[^>]+>", "\n", t)
    t = re.sub(r"&[a-z#0-9]+;", " ", t)
    lines = [re.sub(r"\s+", " ", x).strip() for x in t.split("\n")]
    lines = [x for x in lines if len(x) > 2]
    return "\n".join(dict.fromkeys(lines))[:cap]


def do_dump(a):
    conn = db.get_conn()
    pri = None
    if a.cohort and os.path.exists(a.cohort):
        import openpyxl
        ws = openpyxl.load_workbook(a.cohort).active
        H = [c.value for c in ws[1]]
        i = H.index("id")
        pri = [r[i].value for r in ws.iter_rows(min_row=2) if r[i].value]
    rows = pending(conn, a.n, pri)
    print(f"### 本批 {len(rows)} 人 (剩余待提取见 status)\n")
    for r in rows:
        txt = snapshot_text(conn, r["id"])
        if not txt:
            continue
        print(f"@@@ {r['id']} | {r['name']} | {r['aff'][:40]}")
        print(txt)
        print()


def do_load(a):
    data = json.load(open(a.file, encoding="utf-8"))
    conn = db.get_conn()
    ok = 0
    for pid, ext in data.items():
        pid = int(pid)
        conn.execute("DELETE FROM extractions WHERE person_id=? AND source='homepage' AND version=?",
                     (pid, EXTRACT_VERSION))
        conn.execute("""INSERT INTO extractions (person_id, source, version, model, json)
                        VALUES (?, 'homepage', ?, ?, ?)""",
                     (pid, EXTRACT_VERSION, MODEL_TAG, json.dumps(ext, ensure_ascii=False)))
        normalize(conn, pid, ext)
        ok += 1
    conn.commit()
    print(f"已入库 {ok} 人")
    left = len(pending(conn, 99999))
    print(f"剩余待提取: {left}")


def do_status(a):
    conn = db.get_conn()
    print(f"剩余待提取: {len(pending(conn, 99999))}")
    n = conn.execute("SELECT COUNT(DISTINCT person_id) FROM extractions WHERE model=?",
                     (MODEL_TAG,)).fetchone()[0]
    print(f"本通道已完成: {n}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    d = sub.add_parser("dump"); d.add_argument("--n", type=int, default=30); d.add_argument("--cohort")
    l = sub.add_parser("load"); l.add_argument("file")
    sub.add_parser("status")
    a = ap.parse_args()
    {"dump": do_dump, "load": do_load, "status": do_status}[a.cmd](a)
