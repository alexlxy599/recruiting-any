"""CVPR 2026 候选人导入：从微信收到的人工整理表导入已确认、已去重的北美华人候选人。

用法:
    python3.12 ingest/import_cvpr.py --dry-run   # 解析并打印样例 + 统计，不写库
    python3.12 ingest/import_cvpr.py             # 正式导入
    python3.12 ingest/import_cvpr.py --with-unverified  # 连「4 待核实」一起导（默认不导）

来源表三张精炼 sheet（每人一行，已去重，已确认）:
    3.1 业界          → sector=industry
    3.2 学界(在读)    → sector=academic（招聘重点，含预计毕业年）
    3.3 学界          → sector=academic（老师/研究员）
「4 待核实」默认跳过（低置信、同名难辨），加 --with-unverified 才导且 notes 标记待核实。

去重: 与库内按 linkedin_url → 姓名+机构 两级匹配，命中则只补空字段、追加 CVPR 发表记录，不覆盖既有数据。
每人登记一条 publications(venue='CVPR', year=2026, source='cvpr_sheet_20260623')，
学术/业界视图按 sector / source_type 过滤，会议筛选条按 publications 的 venue 归一。
个人主页 / LinkedIn 列是 =HYPERLINK() 公式，需抽取真实 URL（显示文字是"主页"/"LinkedIn"）。
"""

import argparse
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import db

XLSX = ("/Users/alex/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/"
        "com.tencent.xinWeChat/2.0b4.0.9/90d9f800d0c30108303ef7966e1dc808/Message/MessageTemp/"
        "912034f63d6acf535dcad68172eb44e0/File/CVPR_2026_20260623.xlsx")
VENUE, YEAR = "CVPR", 2026
PUB_SOURCE = "cvpr_sheet_20260623"
NOTE_TAG = "cvpr2026"

# (sheet 名, sector)
CONFIRMED_SHEETS = [("3.1 业界", "industry"), ("3.2 学界(在读)", "academic"), ("3.3 学界", "academic")]
HEADER_ROW = 3  # 1-indexed：第 3 行是真表头（前两行是标题 + 说明）

_HYPERLINK_RE = re.compile(r'=HYPERLINK\("([^"]+)"', re.I)
_YEAR_RE = re.compile(r"(20\d{2})")


def cell_text(v) -> str:
    """单元格 → 文本；若是 =HYPERLINK() 公式则抽 URL。"""
    if v is None:
        return ""
    s = str(v).strip()
    m = _HYPERLINK_RE.search(s)
    if m:
        return m.group(1).strip()
    return s


def norm_url(u: str) -> str | None:
    u = (u or "").strip().rstrip("/")
    return u.lower() if u else None


def split_name(name: str) -> tuple[str, str]:
    parts = (name or "").strip().split()
    if len(parts) <= 1:
        return name.strip(), ""
    return " ".join(parts[:-1]), parts[-1]


def parse_year(s: str):
    m = _YEAR_RE.search(s or "")
    return int(m.group(1)) if m else None


def parse_sheet(wb, sheet_name: str, sector: str, unverified: bool = False) -> list[dict]:
    ws = wb[sheet_name]
    hdr = {h: i for i, h in enumerate(c.value for c in ws[HEADER_ROW]) if h}

    def col(row, name):
        i = hdr.get(name)
        return cell_text(row[i].value) if i is not None and i < len(row) else ""

    people = []
    for row in ws.iter_rows(min_row=HEADER_ROW + 1):
        name = col(row, "姓名")
        if not name:
            continue
        first, last = split_name(name)
        org = col(row, "学校 / 公司")
        tags = [t for t in (col(row, "技术标签1"), col(row, "技术标签2"), col(row, "技术标签3")) if t]
        edu_school = col(row, "毕业院校")
        edu_year = parse_year(col(row, "毕业年份"))
        people.append({
            "first_name": first, "last_name": last,
            "linkedin_url": norm_url(col(row, "LinkedIn")),
            "email": col(row, "邮箱") or None,
            "personal_page": col(row, "个人主页") or None,
            "title": col(row, "职位") or None,
            "institution": org if sector == "academic" else None,
            "company": org if sector == "industry" else None,
            "location": col(row, "国家 / 地区") or None,
            "research_area": ", ".join(tags) or None,
            "tags": tags,
            "expected_graduation": parse_year(col(row, "预计毕业")),
            "sector": sector,
            "source_type": sector,
            "notes": (col(row, "依据 / 备注") or None) if not unverified else
                     f"待核实｜{col(row, '依据 / 备注')}".strip("｜"),
            "edu_school": edu_school or None,
            "edu_year": edu_year,
            "papers": [{"title": col(row, "论文")[:300], "is_first_author": 0}] if col(row, "论文") else [],
        })
    return people


def existing_person_id(conn, p: dict) -> int | None:
    if p["linkedin_url"]:
        row = conn.execute("SELECT id FROM people WHERE LOWER(linkedin_url) = ?",
                           (p["linkedin_url"],)).fetchone()
        if row:
            return row["id"]
    # email 是强独立锚点：机构写法不一致也能命中同一人，降低重复
    if p["email"] and "@" in p["email"]:
        row = conn.execute("SELECT id FROM people WHERE LOWER(email) = ?",
                           (p["email"].lower(),)).fetchone()
        if row:
            return row["id"]
    # 姓名 + 机构（institution 或 company）精确匹配，避免误并同名不同机构
    org = p["institution"] or p["company"]
    if org:
        row = conn.execute(
            """SELECT id FROM people WHERE first_name = ? AND last_name = ?
               AND (institution = ? OR company = ?)""",
            (p["first_name"], p["last_name"], org, org),
        ).fetchone()
        if row:
            return row["id"]
    return None


def upsert(conn, p: dict) -> str:
    pid = existing_person_id(conn, p)
    if pid:
        conn.execute(
            """UPDATE people SET
                 linkedin_url = COALESCE(linkedin_url, ?),
                 email = COALESCE(email, ?),
                 personal_page = COALESCE(personal_page, ?),
                 title = COALESCE(title, ?),
                 institution = COALESCE(institution, ?),
                 company = COALESCE(company, ?),
                 location = COALESCE(location, ?),
                 research_area = COALESCE(research_area, ?),
                 expected_graduation = COALESCE(expected_graduation, ?),
                 sector = COALESCE(sector, ?),
                 source_type = COALESCE(source_type, ?),
                 updated_at = CURRENT_TIMESTAMP
               WHERE id = ?""",
            (p["linkedin_url"], p["email"], p["personal_page"], p["title"],
             p["institution"], p["company"], p["location"], p["research_area"],
             p["expected_graduation"], p["sector"], p["source_type"], pid),
        )
        action = "merged"
    else:
        cur = conn.execute(
            """INSERT INTO people (first_name, last_name, linkedin_url, email, personal_page,
                                   title, institution, company, location, research_area,
                                   expected_graduation, sector, source_type, status, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'new', ?)""",
            (p["first_name"], p["last_name"], p["linkedin_url"], p["email"], p["personal_page"],
             p["title"], p["institution"], p["company"], p["location"], p["research_area"],
             p["expected_graduation"], p["sector"], p["source_type"], p["notes"]),
        )
        pid = cur.lastrowid
        action = "added"
        # 教育（仅新人插，避免覆盖已有）
        if p["edu_school"]:
            conn.execute(
                "INSERT INTO educations (person_id, school, end_year, source) VALUES (?, ?, ?, ?)",
                (pid, p["edu_school"], p["edu_year"], PUB_SOURCE),
            )

    # CVPR 发表记录（按 person+venue+year+title 去重）
    for paper in p["papers"]:
        dup = conn.execute(
            "SELECT 1 FROM publications WHERE person_id = ? AND venue = ? AND year = ? AND title = ?",
            (pid, VENUE, YEAR, paper["title"]),
        ).fetchone()
        if not dup:
            conn.execute(
                """INSERT INTO publications (person_id, venue, year, title, is_first_author, source)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (pid, VENUE, YEAR, paper["title"], paper["is_first_author"], PUB_SOURCE),
            )

    # 标签：技术方向 + 会议标记
    for t in p["tags"]:
        db.add_person_tag(pid, t, category="domain", source=NOTE_TAG, conn=conn)
    db.add_person_tag(pid, f"{VENUE} {YEAR}", category="conference", source=NOTE_TAG, conn=conn)
    return action


def main():
    global XLSX, VENUE, YEAR, PUB_SOURCE, NOTE_TAG
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--with-unverified", action="store_true", help="连「4 待核实」一起导")
    ap.add_argument("--xlsx", default=XLSX, help="来源 xlsx 路径（默认 CVPR 那份）")
    ap.add_argument("--venue", default=VENUE, help="会议缩写，如 ICML / CVPR")
    ap.add_argument("--year", type=int, default=YEAR)
    ap.add_argument("--pub-source", default=None, help="publications.source 标记，默认 {venue}_sheet_20260623")
    ap.add_argument("--note-tag", default=None, help="批次标记，默认 {venue 小写}2026")
    args = ap.parse_args()

    XLSX, VENUE, YEAR = args.xlsx, args.venue, args.year
    PUB_SOURCE = args.pub_source or f"{VENUE.lower()}_sheet_20260623"
    NOTE_TAG = args.note_tag or f"{VENUE.lower()}2026"
    print(f"来源: {os.path.basename(XLSX)} | venue={VENUE} {YEAR} | pub_source={PUB_SOURCE}\n")

    wb = openpyxl.load_workbook(XLSX, data_only=False)
    people = []
    for sheet, sector in CONFIRMED_SHEETS:
        rows = parse_sheet(wb, sheet, sector)
        people.extend(rows)
        print(f"  {sheet:14} → {len(rows)} 人 (sector={sector})")
    if args.with_unverified:
        rows = parse_sheet(wb, "4 待核实", "academic", unverified=True)
        people.extend(rows)
        print(f"  4 待核实        → {len(rows)} 人 (标记待核实)")

    sectors, n_li, n_email, n_page, n_papers = {}, 0, 0, 0, 0
    for p in people:
        sectors[p["sector"]] = sectors.get(p["sector"], 0) + 1
        n_li += bool(p["linkedin_url"]); n_email += bool(p["email"])
        n_page += bool(p["personal_page"]); n_papers += len(p["papers"])
    print(f"\n解析合计: {len(people)} 人 | 界别 {sectors} | 有LinkedIn {n_li} | 有邮箱 {n_email} | "
          f"有主页 {n_page} | {VENUE} 论文 {n_papers} 条")

    print("\n== 样例（前 3 人）==")
    for p in people[:3]:
        print(f"  {p['first_name']} {p['last_name']} | {p['source_type']} | "
              f"机构={p['institution'] or p['company']!r} | 职位={p['title']!r}")
        print(f"     email={p['email']!r} linkedin={p['linkedin_url']!r} page={p['personal_page']!r}")
        print(f"     方向={p['research_area']!r} 预计毕业={p['expected_graduation']} "
              f"毕业院校={p['edu_school']!r}")
        print(f"     论文={p['papers'][0]['title'][:60] if p['papers'] else '(无)'!r}")

    if args.dry_run:
        print("\n[dry-run] 未写库")
        return

    db.init_db()
    conn = db.get_conn()
    stats = {"added": 0, "merged": 0}
    for p in people:
        stats[upsert(conn, p)] += 1
    conn.commit()
    conn.close()
    print(f"\n导入完成: 新增 {stats['added']} 人, 合并进已有档案 {stats['merged']} 人")
    print("重建全文索引...")
    db.rebuild_fts()
    print("done")


if __name__ == "__main__":
    main()
